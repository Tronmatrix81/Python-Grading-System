import os
import openpyxl

absolutePath=os.path.dirname(os.path.abspath(__file__)) # Find actual directory
excelPath=os.path.join(absolutePath, "queso.xlsx") # Find directory's excel
idColumn = 1
nameColumn = 2

def open_excel():
    global excelFile, excelPath

    try:
        excelFile=openpyxl.load_workbook(excelPath)
        
    except:
        print("Error: File doesn't exist or is already open.")

    return excelFile

def read_cell(sheet:str, row:int, column:int):
        excelFile=open_excel()
        excelSheet=excelFile[sheet]

        return excelSheet.cell(row, column).value

def write_cell(row:float, column:float, data:str):
    global excelPath

    excelFile=open_excel()#abre el archivo
    excelSheet=excelFile.active
    excelSheet.cell(row, column, data)
    #celda segun row columna y el dato, ese es el orden pues

    try:
        excelFile.save(excelPath)
    except PermissionError:
        print("Error: Excel file is currently open.")

def add_member_ui():
    print("~~Add member~~")

    name = input("Type name: ")
    id = input("Type matricula: ")

    add_member(name, id)
   

def add_member(name: str, id: str):
    global excelPath

    excelFile=open_excel()
    excelSheet=excelFile["promedio"]
    #maxColumns = excelSheet.max_column
    maxRows = excelSheet.max_row

    numberOfStudents = maxRows - 1 #We remove the first line cause its for headers

    if numberOfStudents < 5:

        for i in range(2, maxRows + 1):
            cellX = excelSheet.cell(row=i, column=nameColumn).value
            if (cellX == name):
                print("Student already registered!")
                os.system("pause")
                return

        excelSheet.cell(row= maxRows + 1, column=nameColumn, value=name) #Assign the values to the cells
        excelSheet.cell(row= maxRows + 1, column=idColumn, value=id) #Assign the values to the cells

        try:
            excelFile.save(excelPath)
        except PermissionError:
            print("Error: Excel file is currently open.")
    else:
        print("Number of students are at maximum capacity.")

    os.system("pause")

def FindStudent(name: str):
    global excelPath

    excelFile=open_excel()
    excelSheet=excelFile["promedio"]
    #maxColumns = excelSheet.max_column
    maxRows = excelSheet.max_row

    for i in range(2, maxRows + 1):
        cellX = excelSheet.cell(i, nameColumn).value
        if (cellX == name):
            id = excelSheet.cell(i, idColumn).value
            return {id: name}
        
#def find_student_ui():


def remove_member_ui():
    os.system("cls")

    print("~~Remove member~~")

    name = input("Type student name: ")
    remove_member(name)
    


def remove_member(name: str):
    global excelPath

    excelFile=open_excel()
    excelSheet=excelFile["promedio"]
    #maxColumns = excelSheet.max_column
    maxRows = excelSheet.max_row

    numberOfStudents = maxRows - 1 #We remove the headers

    lineToRemove = -1 #default value 

    if (numberOfStudents == 1):
        print("Cannot remove any more students! At least one student is required.")
    else:
        for i in range(2, maxRows + 1):
            cellX = excelSheet.cell(row=i, column=nameColumn).value

            if (cellX == name):
                lineToRemove = i
        
        if (lineToRemove != -1):
            excelSheet.delete_rows(lineToRemove, 1)

            try:
                excelFile.save(excelPath)
            except PermissionError:
                print("Error: Excel file is currently open.")
        else:
            print("Student not found!")

    os.system("pause")    